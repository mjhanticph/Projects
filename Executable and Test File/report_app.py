import pandas as pd
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import datetime

df = pd.read_excel('test.xlsx')

df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

df = df.drop(columns=['APPROVE DATE', 'DATE DENIED','PROCESSED DATE','PAYORLINK NO.',
                      'TAX RATE','TAX AMOUNT','PAYABLE AMOUNT','GGD REFERENCE NO',
                     'MEMBER EXCESS','DATE PAID','CHECK NUMBER','DATE OF POSTING',
                     'APPROVER','PROCESSOR','PLAN TYPE'])

df['Date'] = df['RECEIVED DATE'].where(df['PAYEE TYPE'] == 'REIMBURSEMENT', df['ILLNESS DATE'])

while True:
    # Prompt the user for the start and end dates
    start_date = input('Enter the start date (YYYY-MM-DD): ')
    end_date = input('Enter the end date (YYYY-MM-DD): ')

    # Convert the start_date and end_date strings to datetime objects
    try:
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        # If the date strings are not in the correct format, print an error message and continue the loop
        print("Invalid date format. Please enter the dates in the YYYY-MM-DD format.")
        continue

    # If the start date is after the end date, print an error message and continue the loop
    if start_date > end_date:
        print("The start date must be before the end date. Please try again.")
        continue

    # If the start and end dates are valid, break out of the loop
    break

# Format the start_date and end_date variables as long dates
start_date_formatted = start_date.strftime('%B %d, %Y')
end_date_formatted = end_date.strftime('%B %d, %Y')

# Concatenate the formatted dates with a 'to' separator
policy_period = f"{start_date_formatted} to {end_date_formatted}"

filtered_df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date) & (df['CLAIM STATUS'] != 'denied')].dropna(subset=['CLAIM AMOUNT'])

### GETTING THE INFORMATION NEEDED FOR THE SUMMARY SHEET
# Company Name
name = filtered_df.iloc[0]['CLIENT NAME']

# Date extracted indicator
# Get the current date and format it as a long date
now = datetime.datetime.now()
date_formatted = now.strftime('%B %d, %Y')
# Concatenate the date with the "Date extracted:" string
date_extracted = f"Date extracted: {date_formatted}"

# SUMMARY OF COST
# Group the DataFrame by CLAIM TYPE and calculate the sum of CLAIM AMOUNT for each group
summary_df = filtered_df.groupby('CLAIM TYPE')['CLAIM AMOUNT'].sum()
# Append a total row to the summary DataFrame
total = summary_df.sum()
summary_df = pd.concat([summary_df, pd.Series(total, index=['Total'])])

### CREATING A NEW WORKBOOK

workbook = op.Workbook()
summary_sheet = workbook.create_sheet("Summary")

workbook.remove(workbook.worksheets[0])

# Set the account name in cells B3 to C3 (merged)
summary_sheet.merge_cells('B3:C3')
summary_sheet['B3'] = name
alignment = Alignment(horizontal='center')
summary_sheet['B3'].alignment = alignment

# Set the date range in cells B4 to C4 (merged)
summary_sheet.merge_cells('B4:C4')
summary_sheet['B4'] = policy_period
alignment = Alignment(horizontal='center')
summary_sheet['B4'].alignment = alignment

# Set the date extracted in cells B5 to C5 (merged)
summary_sheet.merge_cells('B5:C5')
summary_sheet['B5'] = date_extracted
alignment = Alignment(horizontal='center')
summary_sheet['B5'].alignment = alignment

# Set the Claim Type and Claim Amount headings in cells B6 and C6
summary_sheet['B6'] = 'CLAIM TYPE'
summary_sheet['C6'] = 'CLAIM AMOUNT'
alignment = Alignment(horizontal='center')
summary_sheet['B6'].alignment = alignment
summary_sheet['B6'].alignment = alignment

# Set the font of the account name, date range, and date extracted cells to be bold
summary_sheet['B3'].font = op.styles.Font(bold=True)
summary_sheet['B4'].font = op.styles.Font(bold=True)
summary_sheet['B5'].font = op.styles.Font(bold=True)

# Set the font of the claim type and claim amount headers to be bold
summary_sheet['B6'].font = op.styles.Font(bold=True)
summary_sheet['C6'].font = op.styles.Font(bold=True)

# Set the number format of the cells containing the claim amounts to be a float with a comma separator
for row in summary_sheet.iter_rows(min_row=7, max_row=summary_df.shape[0]+6, min_col=2, max_col=3):
    for cell in row:
        cell.number_format = '#,##0.00'

# Loop through the index labels in the summary_df dataframe and add the data to the summary sheet
for i, label in enumerate(summary_df.index):
    summary_sheet.cell(row=i+7, column=2).value = label
    summary_sheet.cell(row=i+7, column=3).value = summary_df[label]

# Iterate through all the columns in the 'summary' sheet and set the column width to the maximum width of any cell in that column
for column_cells in summary_sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    summary_sheet.column_dimensions[column_cells[0].column_letter].width = length

raw_sheet = workbook.create_sheet(title='raw')
for r in dataframe_to_rows(filtered_df, index=False, header=True):
    raw_sheet.append(r)

for column_cells in raw_sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    raw_sheet.column_dimensions[column_cells[0].column_letter].width = length



# Convert the datetime object to a string in the desired format
start_date_str = start_date.strftime('%m%d%y')
end_date_str = end_date.strftime('%m%d%y')

date_string = start_date_str+'-'+end_date_str
# print(date_string)

workbook.save(name +'_'+'CLAIMS UTIL'+'_'+ date_string + '.xlsx')


## Overview
# print(name)
# print(policy_period)
# print(date_extracted)
# print(summary_df)
# print(filtered_df)
